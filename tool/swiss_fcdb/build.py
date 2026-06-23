#!/usr/bin/env python3
"""Build assets/swiss_foods.csv.gz from the Swiss Food Composition Database.

Source: Swiss Food Composition Database (naehrwertdaten.ch), published by the
Federal Food Safety and Veterinary Office (FSVO/BLV). Free for commercial use
"subject to acknowledgment of the source" — Knabberfuchs credits it in
Settings → About.

The DB ships one Excel file per language (DE/FR/IT/EN), all sharing a stable
numeric `ID` column. We join the "generic foods" sheet on that ID to produce a
single multilingual row per food. English is the canonical/fallback `name`;
DE/FR/IT are stored as overrides. A `search_text` field concatenates every
language's name + synonyms so search works regardless of UI language.

Place the per-language .xlsx files next to this script as de.xlsx / fr.xlsx /
it.xlsx / en.xlsx (any missing language is simply left blank). Then run:
    python3 tool/swiss_fcdb/build.py
"""
import csv
import gzip
import io
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "..", "..", "assets", "swiss_foods.csv.gz")
PORTIONS = os.path.join(HERE, "..", "portions", "portions.csv")


def _load_portions():
    """Curated natural-portion weights, keyed by exact English name (lower-cased).

    Returns {name_lower: (unit_key, grams_str)}. See tool/portions/.
    """
    out = {}
    if not os.path.exists(PORTIONS):
        return out
    with open(PORTIONS, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[row["name_en"].strip().lower()] = (
                row["unit"].strip(), row["grams"].strip())
    return out

# The "generic foods" sheet is named per language; match by position/keywords.
GENERIC_SHEET = {
    "de": "Generische Lebensmittel",
    "en": "Generic Foods",
    "fr": "Aliments génériques",
    "it": "Alimenti generici",  # best guess; corrected when it.xlsx arrives
}

# Column indices (0-based) in the generic sheet — identical across languages.
C_ID = 0
C_NAME = 3
C_SYN = 4
C_KCAL = 11
C_FAT = 14
C_SATFAT = 17
C_CARB = 29
C_SUGAR = 32
C_FIBRE = 38
C_PROTEIN = 41
C_SALT = 44
C_SODIUM = 101
HEADER_ROW = 2  # rows 0-1 are title/notes; row 2 is the header; data from row 3


def _num(v):
    """Parse a Swiss FCDB nutrient cell → float or None.

    Numbers pass through; "Sp."/"tr." (trace) → 0; "k.A."/"n.v."/blank → None.
    """
    if v is None:
        return None
    s = str(v).strip().replace(",", ".")
    if s == "":
        return None
    if s.lower() in ("sp.", "tr.", "trace", "spur"):
        return 0.0
    m = re.match(r"^-?\d+(\.\d+)?$", s)
    return float(s) if m else None


def _open_sheet(lang):
    path = os.path.join(HERE, f"{lang}.xlsx")
    if not os.path.exists(path):
        return None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = GENERIC_SHEET.get(lang)
    if name not in wb.sheetnames:
        # fall back to the first sheet
        name = wb.sheetnames[0]
    return list(wb[name].iter_rows(values_only=True))


def _names(rows):
    """ID -> (name, synonyms) for one language file."""
    out = {}
    if not rows:
        return out
    for r in rows[HEADER_ROW + 1:]:
        if not r or r[C_ID] is None:
            continue
        out[str(r[C_ID]).strip()] = (
            (str(r[C_NAME]).strip() if r[C_NAME] else ""),
            (str(r[C_SYN]).strip() if len(r) > C_SYN and r[C_SYN] else ""),
        )
    return out


def main():
    langs = {l: _open_sheet(l) for l in ("en", "de", "fr", "it")}
    present = [l for l, r in langs.items() if r]
    print(f"languages present: {', '.join(present) or '(none)'}", file=sys.stderr)
    if "en" not in present:
        sys.exit("en.xlsx is required as the canonical/fallback language.")

    names = {l: _names(r) for l, r in langs.items()}
    base = langs["en"]  # nutrients read from EN (identical across languages)
    portions = _load_portions()
    matched = 0

    rows_out = []
    for r in base[HEADER_ROW + 1:]:
        if not r or r[C_ID] is None:
            continue
        fid = str(r[C_ID]).strip()
        kcal = _num(r[C_KCAL])
        if kcal is None:
            continue
        name_en = names["en"].get(fid, ("", ""))[0]
        if not name_en:
            continue
        name_de = names["de"].get(fid, ("", ""))[0]
        name_fr = names["fr"].get(fid, ("", ""))[0]
        name_it = names["it"].get(fid, ("", ""))[0]

        # search_text: every language's name + synonyms, lower-cased, deduped.
        bits = []
        for l in ("en", "de", "fr", "it"):
            nm, syn = names[l].get(fid, ("", ""))
            bits.append(nm)
            bits.append(syn)
        toks = []
        seen = set()
        for tok in re.split(r"[^0-9a-zA-Zäöüàâçéèêëîïôûùœ]+", " ".join(bits).lower()):
            if tok and tok not in seen:
                seen.add(tok)
                toks.append(tok)
        search_text = " ".join(toks)

        serving_unit, serving_g = portions.get(name_en.strip().lower(), ("", ""))
        if serving_g:
            matched += 1

        rows_out.append([
            fid, name_en, name_de, name_fr, name_it,
            _fmt(kcal), _fmt(_num(r[C_PROTEIN])), _fmt(_num(r[C_CARB])),
            _fmt(_num(r[C_FAT])), _fmt(_num(r[C_FIBRE])), _fmt(_num(r[C_SUGAR])),
            _fmt(_num(r[C_SATFAT])), _fmt(_num(r[C_SODIUM])),
            search_text, serving_g, serving_unit,
        ])

    rows_out.sort(key=lambda x: x[1].lower())  # by English name for stable diffs
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "id", "name_en", "name_de", "name_fr", "name_it",
        "kcal100", "protein100", "carb100", "fat100", "fiber100", "sugar100",
        "satfat100", "sodium_mg100", "search_text", "serving_g", "serving_unit",
    ])
    w.writerows(rows_out)
    raw = buf.getvalue().encode("utf-8")
    with gzip.open(OUT, "wb", compresslevel=9) as f:
        f.write(raw)
    print(f"wrote {len(rows_out)} foods ({matched} with a natural portion) -> "
          f"{os.path.relpath(OUT, os.path.join(HERE,'..','..'))} "
          f"({os.path.getsize(OUT)//1024} KB gz, {len(raw)//1024} KB raw)", file=sys.stderr)


def _fmt(v):
    if v is None:
        return ""
    # trim trailing .0 for compactness
    return str(int(v)) if v == int(v) else f"{v:g}"


if __name__ == "__main__":
    main()
